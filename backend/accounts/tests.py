from io import BytesIO

from django.test import TestCase
from openpyxl import Workbook

from academics.models import FacultySubject, Subject
from audit.models import AuditLog
from core.testing import (
    INITIAL, STRONG, bearer, client_for, login, make_admin, make_faculty, make_student, make_subject,
)

from .models import Role, User


class AuthenticationTests(TestCase):
    def setUp(self):
        self.admin = make_admin(email="admin@example.edu")
        self.faculty = make_faculty(email="prof@example.edu")
        self.student = make_student(email="stu@example.edu")

    def test_each_role_can_log_in_on_its_own_endpoint(self):
        for role, user in (("admin", self.admin), ("faculty", self.faculty), ("student", self.student)):
            res = login(client_for(), role, user.email, STRONG)
            self.assertEqual(res.status_code, 200, res.content)
            self.assertIn("access", res.data)
            self.assertEqual(res.data["user"]["role"], role)
            self.assertFalse(res.data["must_change_password"])

    def test_wrong_role_endpoint_is_rejected_like_bad_password(self):
        res = login(client_for(), "admin", self.student.email, STRONG)
        self.assertEqual(res.status_code, 401)
        self.assertEqual(res.data["error"]["code"], "INVALID_CREDENTIALS")

    def test_invalid_password(self):
        res = login(client_for(), "student", self.student.email, "nope")
        self.assertEqual(res.status_code, 401)
        self.assertTrue(AuditLog.objects.filter(action="auth.login_failed").exists())

    def test_inactive_user_cannot_log_in(self):
        self.student.status = "discontinued"
        self.student.save()
        res = login(client_for(), "student", self.student.email, STRONG)
        self.assertEqual(res.status_code, 401)

    def test_unauthenticated_request_gets_structured_error(self):
        res = client_for().get("/api/auth/me/")
        self.assertEqual(res.status_code, 401)
        self.assertEqual(res.data["error"]["code"], "AUTHENTICATION_REQUIRED")

    def test_token_round_trip_and_refresh_rotation(self):
        res = login(client_for(), "faculty", self.faculty.email, STRONG)
        client = bearer(client_for(), res.data["access"])
        self.assertEqual(client.get("/api/auth/me/").status_code, 200)
        refreshed = client_for().post("/api/auth/refresh/", {"refresh": res.data["refresh"]}, format="json")
        self.assertEqual(refreshed.status_code, 200)
        reused = client_for().post("/api/auth/refresh/", {"refresh": res.data["refresh"]}, format="json")
        self.assertEqual(reused.status_code, 401)

    def test_login_opens_application_session_and_logout_closes_it(self):
        res = login(client_for(), "student", self.student.email, STRONG)
        session_id = res.data["session_id"]
        client = bearer(client_for(), res.data["access"])
        out = client.post("/api/auth/logout/", {"refresh": res.data["refresh"], "session_id": session_id}, format="json")
        self.assertEqual(out.status_code, 204)
        from activity.models import ApplicationSession
        session = ApplicationSession.objects.get(pk=session_id)
        self.assertIsNotNone(session.logout_at)
        self.assertEqual(session.ended_by, "logout")


class MandatoryPasswordResetTests(TestCase):
    def setUp(self):
        self.student = make_student(email="new@example.edu", password=INITIAL, must_change=True)

    def test_flagged_user_is_blocked_everywhere_except_password_change(self):
        res = login(client_for(), "student", self.student.email, INITIAL)
        self.assertTrue(res.data["must_change_password"])
        client = bearer(client_for(), res.data["access"])
        blocked = client.get("/api/student/subjects/")
        self.assertEqual(blocked.status_code, 403)
        self.assertEqual(blocked.data["error"]["code"], "PASSWORD_CHANGE_REQUIRED")
        self.assertEqual(client.get("/api/auth/me/").status_code, 200)

    def test_password_change_lifts_block_and_is_audited(self):
        res = login(client_for(), "student", self.student.email, INITIAL)
        client = bearer(client_for(), res.data["access"])
        weak = client.post("/api/auth/password/change/", {"current_password": INITIAL, "new_password": "short"}, format="json")
        self.assertEqual(weak.status_code, 400)
        changed = client.post("/api/auth/password/change/", {"current_password": INITIAL, "new_password": STRONG}, format="json")
        self.assertEqual(changed.status_code, 200, changed.content)
        self.assertFalse(changed.data["must_change_password"])
        client = bearer(client_for(), changed.data["access"])
        self.assertEqual(client.get("/api/student/subjects/").status_code, 200)
        self.student.refresh_from_db()
        self.assertFalse(self.student.must_change_password)
        self.assertTrue(AuditLog.objects.filter(action="user.password_changed", target_id=str(self.student.id)).exists())

    def test_wrong_current_password_rejected(self):
        res = login(client_for(), "student", self.student.email, INITIAL)
        client = bearer(client_for(), res.data["access"])
        bad = client.post("/api/auth/password/change/", {"current_password": "wrong", "new_password": STRONG}, format="json")
        self.assertEqual(bad.status_code, 400)
        self.assertEqual(bad.data["error"]["code"], "INVALID_CURRENT_PASSWORD")


class RoleIsolationTests(TestCase):
    def setUp(self):
        self.admin = make_admin()
        self.faculty = make_faculty()
        self.student = make_student()

    def test_student_cannot_call_admin_apis(self):
        res = client_for(self.student).get("/api/admin/students/")
        self.assertEqual(res.status_code, 403)
        res = client_for(self.student).post("/api/admin/subjects/", {"name": "X", "code": "X1"}, format="json")
        self.assertEqual(res.status_code, 403)

    def test_faculty_cannot_call_admin_apis(self):
        self.assertEqual(client_for(self.faculty).get("/api/admin/faculty/").status_code, 403)
        self.assertEqual(client_for(self.faculty).get("/api/admin/audit-logs/").status_code, 403)

    def test_student_cannot_call_faculty_apis(self):
        self.assertEqual(client_for(self.student).get("/api/faculty/subjects/").status_code, 403)

    def test_faculty_cannot_call_student_apis(self):
        self.assertEqual(client_for(self.faculty).get("/api/student/subjects/").status_code, 403)


class AdminUserManagementTests(TestCase):
    def setUp(self):
        self.admin = make_admin()
        self.client = client_for(self.admin)
        self.subject = make_subject()

    def test_manual_faculty_creation_sets_initial_password_and_subjects(self):
        res = self.client.post("/api/admin/faculty/", {
            "email": "New.Prof@Example.edu", "full_name": "New Prof",
            "profile": {"employee_id": "E9", "department": "CS"},
            "subject_ids": [str(self.subject.id)],
        }, format="json")
        self.assertEqual(res.status_code, 201, res.content)
        user = User.objects.get(email="new.prof@example.edu")
        self.assertEqual(user.role, Role.FACULTY)
        self.assertTrue(user.must_change_password)
        self.assertTrue(user.check_password(INITIAL))
        self.assertEqual(user.faculty_profile.employee_id, "E9")
        self.assertTrue(FacultySubject.objects.filter(faculty=user, subject=self.subject, status="active").exists())
        self.assertTrue(AuditLog.objects.filter(action="user.created", target_id=str(user.id)).exists())

    def test_duplicate_email_conflicts(self):
        self.client.post("/api/admin/students/", {"email": "s@example.edu", "full_name": "S"}, format="json")
        res = self.client.post("/api/admin/students/", {"email": "S@example.edu", "full_name": "S2"}, format="json")
        self.assertEqual(res.status_code, 409)
        self.assertEqual(res.data["error"]["code"], "USER_EXISTS")

    def test_students_endpoint_only_creates_students(self):
        res = self.client.post("/api/admin/students/", {"email": "s@example.edu", "full_name": "S", "profile": {"roll_number": "R1"}}, format="json")
        self.assertEqual(res.status_code, 201)
        self.assertEqual(res.data["role"], "student")
        self.assertEqual(res.data["profile"]["roll_number"], "R1")

    def test_discontinue_reactivate_and_history_preserved(self):
        student = make_student()
        make_subject(code="CS102")
        res = self.client.post(f"/api/admin/students/{student.id}/discontinue/", {"reason": "left"}, format="json")
        self.assertEqual(res.status_code, 200)
        self.assertEqual(res.data["status"], "discontinued")
        self.assertTrue(User.objects.filter(pk=student.pk).exists())
        self.assertEqual(self.client.post(f"/api/admin/students/{student.id}/discontinue/", {}, format="json").status_code, 409)
        res = self.client.post(f"/api/admin/students/{student.id}/reactivate/", {}, format="json")
        self.assertEqual(res.data["status"], "active")

    def test_delete_student_removes_the_account_and_its_records(self):
        """The People screen deletes accounts now, so the row and everything
        hanging off it must go while the audit trail survives."""
        from academics.models import Enrollment

        student = make_student()
        subject = make_subject(code="CS103")
        Enrollment.objects.create(student=student, subject=subject)

        res = self.client.delete(f"/api/admin/students/{student.id}/", {"reason": "left the programme"}, format="json")

        self.assertEqual(res.status_code, 200, res.content)
        self.assertFalse(User.objects.filter(pk=student.pk).exists())
        self.assertFalse(Enrollment.objects.filter(student_id=student.pk).exists())
        entry = AuditLog.objects.get(action="user.deleted")
        self.assertEqual(entry.summary["reason"], "left the programme")

    def test_delete_faculty_releases_its_subject_assignments(self):
        faculty = make_faculty()
        subject = make_subject(code="CS104")
        FacultySubject.objects.create(faculty=faculty, subject=subject)
        res = self.client.delete(f"/api/admin/faculty/{faculty.id}/")
        self.assertEqual(res.status_code, 200, res.content)
        self.assertFalse(User.objects.filter(pk=faculty.pk).exists())
        self.assertFalse(FacultySubject.objects.filter(faculty_id=faculty.pk).exists())
        self.assertTrue(Subject.objects.filter(pk=subject.pk).exists())

    def test_cannot_delete_own_account(self):
        admin = make_admin(email="second-admin@example.edu")
        res = client_for(admin).delete(f"/api/admin/faculty/{admin.id}/")
        self.assertEqual(res.status_code, 404)  # an admin is not inside the faculty scope
        self.assertTrue(User.objects.filter(pk=admin.pk).exists())

    def test_cannot_discontinue_self(self):
        res = self.client.post(f"/api/admin/faculty/{self.admin.id}/discontinue/", {}, format="json")
        self.assertEqual(res.status_code, 404)  # admin is not in the faculty scope

    def test_unknown_id_is_404_not_latest(self):
        make_student()
        res = self.client.get("/api/admin/students/00000000-0000-0000-0000-000000000000/")
        self.assertEqual(res.status_code, 404)
        self.assertEqual(res.data["error"]["code"], "NOT_FOUND")

    def test_admin_password_reset_forces_change(self):
        student = make_student()
        res = self.client.post(f"/api/admin/students/{student.id}/reset-password/", {}, format="json")
        self.assertEqual(res.status_code, 200)
        student.refresh_from_db()
        self.assertTrue(student.must_change_password)
        self.assertTrue(student.check_password(INITIAL))


def _workbook(headers, rows):
    wb = Workbook()
    ws = wb.active
    ws.append(headers)
    for row in rows:
        ws.append(row)
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    buf.name = "users.xlsx"
    return buf


class ExcelImportTests(TestCase):
    def setUp(self):
        self.admin = make_admin()
        self.client = client_for(self.admin)
        make_student(email="exists@example.edu")
        self.subject = make_subject(code="CS101")

    def test_student_import_reports_created_existing_and_invalid(self):
        wb = _workbook(["Name", "Email", "Roll Number"], [
            ["Alice", "alice@example.edu", "R1"],
            ["Bob", "exists@example.edu", "R2"],
            ["", "noname@example.edu", "R3"],
            ["Dana", "not-an-email", "R4"],
            ["Eve", "alice@example.edu", "R5"],
            ["Frank", "frank@example.edu", "R6"],
        ])
        res = self.client.post("/api/admin/students/import/", {"file": wb}, format="multipart")
        self.assertEqual(res.status_code, 200, res.content)
        self.assertEqual(res.data["total_rows"], 6)
        self.assertEqual(res.data["created"], 2)
        self.assertEqual(res.data["already_existing"], 1)
        self.assertEqual(res.data["invalid"], 3)
        rows_with_errors = {e["row"] for e in res.data["errors"]}
        self.assertEqual(rows_with_errors, {3, 4, 5, 6})
        alice = User.objects.get(email="alice@example.edu")
        self.assertTrue(alice.must_change_password)
        self.assertEqual(alice.student_profile.roll_number, "R1")

    def test_faculty_import_resolves_subject_codes(self):
        wb = _workbook(["name", "email", "subject_codes", "department"], [
            ["Prof A", "a@example.edu", "cs101", "CS"],
            ["Prof B", "b@example.edu", "CS101, NOPE", "CS"],
        ])
        res = self.client.post("/api/admin/faculty/import/", {"file": wb}, format="multipart")
        self.assertEqual(res.data["created"], 1)
        self.assertEqual(res.data["invalid"], 1)
        a = User.objects.get(email="a@example.edu")
        self.assertTrue(FacultySubject.objects.filter(faculty=a, subject=self.subject).exists())
        self.assertFalse(User.objects.filter(email="b@example.edu").exists())

    def test_missing_headers_rejected(self):
        wb = _workbook(["Name", "Phone"], [["X", "1"]])
        res = self.client.post("/api/admin/students/import/", {"file": wb}, format="multipart")
        self.assertEqual(res.status_code, 400)
        self.assertEqual(res.data["error"]["code"], "MISSING_HEADERS")

    def test_non_xlsx_rejected(self):
        buf = BytesIO(b"name,email\n")
        buf.name = "users.csv"
        res = self.client.post("/api/admin/students/import/", {"file": buf}, format="multipart")
        self.assertEqual(res.status_code, 400)


class ImportHeaderNormalisationTests(TestCase):
    def test_documented_header_variants_are_accepted(self):
        from accounts.services.excel_import import _normalize_header
        cases = {"Full Name": "name", "E-mail": "email", "E-mail Address": "email", "Email": "email",
                 "Roll No": "roll_number", "Roll No.": "roll_number", "Roll Number": "roll_number",
                 "Employee ID": "employee_id", "Subjects": "subject_codes", "Subject Codes": "subject_codes",
                 "Batch": "batch", "Phone": "phone", "Department": "department"}
        for raw, expected in cases.items():
            self.assertEqual(_normalize_header(raw), expected, raw)


class ImportTemplateTests(TestCase):
    """The screen, the template and the parser must describe one sheet."""

    def setUp(self):
        self.client = client_for(make_admin())

    def test_template_headers_match_what_the_parser_accepts(self):
        import base64
        from io import BytesIO

        from openpyxl import load_workbook

        from accounts.services.excel_import import OPTIONAL_HEADERS, REQUIRED_HEADERS

        res = self.client.get("/api/admin/students/import/template/")
        self.assertEqual(res.status_code, 200, res.content)
        names = [c["name"] for c in res.data["columns"]]
        self.assertEqual(set(names), REQUIRED_HEADERS | OPTIONAL_HEADERS[Role.STUDENT])
        self.assertTrue(all(c["required"] for c in res.data["columns"] if c["name"] in REQUIRED_HEADERS))

        book = load_workbook(BytesIO(base64.b64decode(res.data["content_base64"])))
        header = [c.value for c in book.active[1]]
        self.assertEqual(header, names)
        self.assertGreaterEqual(book.active.max_row, 2, "the template carries example rows")

    def test_the_template_it_serves_imports_cleanly(self):
        import base64
        from io import BytesIO

        res = self.client.get("/api/admin/students/import/template/")
        payload = BytesIO(base64.b64decode(res.data["content_base64"]))
        payload.name = "template.xlsx"

        imported = self.client.post("/api/admin/students/import/", {"file": payload}, format="multipart")

        self.assertEqual(imported.status_code, 200, imported.content)
        self.assertEqual(imported.data["invalid"], 0, imported.data["errors"])
        self.assertEqual(imported.data["created"], 2)

    def test_faculty_template_covers_subject_codes(self):
        res = self.client.get("/api/admin/faculty/import/template/")
        self.assertIn("subject_codes", [c["name"] for c in res.data["columns"]])
        aliases = {c["name"]: c["aliases"] for c in res.data["columns"]}
        self.assertIn("full_name", aliases["name"])
