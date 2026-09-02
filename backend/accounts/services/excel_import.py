"""Excel bulk import.

Parsing (this module) is deliberately separate from user creation
(accounts.services.users). The parser turns a workbook into row dicts and
row-level errors; the importer feeds valid rows through create_user with a
savepoint per row so one bad row never poisons the batch.
"""
import re
from dataclasses import dataclass, field
from io import BytesIO

from django.db import transaction

from audit import services as audit
from core.exceptions import APIError, Conflict, ValidationFailed

from ..models import Role, User
from .users import NewUser, create_user, normalize_email

REQUIRED_HEADERS = {"name", "email"}
OPTIONAL_HEADERS = {
    Role.FACULTY: {"employee_id", "department", "designation", "phone", "subject_codes"},
    Role.STUDENT: {"roll_number", "program", "batch", "phone"},
}
# Keys are in canonical form: lower-case, with spaces, hyphens and dots
# collapsed to single underscores (see _normalize_header).
HEADER_ALIASES = {
    "full_name": "name", "student_name": "name", "faculty_name": "name",
    "e_mail": "email", "email_address": "email", "e_mail_address": "email", "mail": "email",
    "roll_no": "roll_number", "roll": "roll_number", "rollno": "roll_number",
    "employee_id": "employee_id", "emp_id": "employee_id", "employee_no": "employee_id",
    "subjects": "subject_codes", "subject_code": "subject_codes", "subject": "subject_codes",
    "programme": "program", "course": "program", "phone_number": "phone", "mobile": "phone",
}
MAX_ROWS = 5000


@dataclass
class ParsedRow:
    row_number: int
    data: dict
    errors: list = field(default_factory=list)


@dataclass
class ImportReport:
    total_rows: int = 0
    created: int = 0
    already_existing: int = 0
    invalid: int = 0
    created_users: list = field(default_factory=list)
    errors: list = field(default_factory=list)

    def as_dict(self):
        return {
            "total_rows": self.total_rows,
            "created": self.created,
            "already_existing": self.already_existing,
            "invalid": self.invalid,
            "created_users": self.created_users,
            "errors": self.errors,
        }


def _normalize_header(value):
    """'E-mail Address' -> 'email', 'Roll No.' -> 'roll_number', 'Batch' -> 'batch'."""
    key = re.sub(r"[\s\-.]+", "_", str(value or "").strip().lower()).strip("_")
    return HEADER_ALIASES.get(key, key)


def parse_workbook(file_obj, role) -> list[ParsedRow]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover
        raise APIError("Excel support is not installed.", code="EXCEL_UNAVAILABLE", status_code=500) from exc

    raw = file_obj.read() if hasattr(file_obj, "read") else file_obj
    try:
        workbook = load_workbook(BytesIO(raw), read_only=True, data_only=True)
    except Exception as exc:
        raise ValidationFailed("The uploaded file is not a readable .xlsx workbook.", code="INVALID_WORKBOOK") from exc

    sheet = workbook.active
    rows = sheet.iter_rows(values_only=True)
    try:
        header_row = next(rows)
    except StopIteration:
        raise ValidationFailed("The workbook is empty.", code="EMPTY_WORKBOOK")

    headers = [_normalize_header(h) for h in header_row]
    missing = REQUIRED_HEADERS - set(headers)
    if missing:
        raise ValidationFailed(
            "Required columns are missing.", code="MISSING_HEADERS",
            details={"missing": sorted(missing), "found": headers},
        )
    allowed = REQUIRED_HEADERS | OPTIONAL_HEADERS.get(role, set())

    parsed = []
    seen_emails = {}
    for offset, values in enumerate(rows, start=2):
        if values is None or all(v in (None, "") for v in values):
            continue
        if len(parsed) >= MAX_ROWS:
            raise ValidationFailed(f"Imports are limited to {MAX_ROWS} rows per file.", code="TOO_MANY_ROWS")
        record = {}
        for header, value in zip(headers, values):
            if header in allowed:
                record[header] = "" if value is None else str(value).strip()
        row = ParsedRow(row_number=offset, data=record)
        email = normalize_email(record.get("email"))
        record["email"] = email
        if not email or "@" not in email:
            row.errors.append("Invalid or missing email.")
        if not record.get("name"):
            row.errors.append("Missing name.")
        if email and email in seen_emails:
            row.errors.append(f"Duplicate of row {seen_emails[email]} in this file.")
        elif email:
            seen_emails[email] = offset
        parsed.append(row)
    return parsed


def import_users(actor, file_obj, role, request=None) -> ImportReport:
    if role not in (Role.FACULTY, Role.STUDENT):
        raise ValidationFailed("Role must be faculty or student.", code="INVALID_ROLE")
    rows = parse_workbook(file_obj, role)
    report = ImportReport(total_rows=len(rows))

    from academics.models import Subject

    for row in rows:
        if row.errors:
            report.invalid += 1
            report.errors.append({"row": row.row_number, "email": row.data.get("email"), "errors": row.errors})
            continue

        data = row.data
        profile = {k: v for k, v in data.items() if k not in ("name", "email", "subject_codes")}
        subject_ids = []
        if role == Role.FACULTY and data.get("subject_codes"):
            codes = [c.strip().upper() for c in data["subject_codes"].split(",") if c.strip()]
            found = {s.code: s.id for s in Subject.objects.filter(code__in=codes)}
            unknown = [c for c in codes if c not in found]
            if unknown:
                report.invalid += 1
                report.errors.append({"row": row.row_number, "email": data["email"],
                                      "errors": [f"Unknown subject codes: {', '.join(unknown)}"]})
                continue
            subject_ids = list(found.values())

        try:
            with transaction.atomic():
                user = create_user(
                    actor,
                    NewUser(email=data["email"], full_name=data["name"], role=role,
                            profile=profile, subject_ids=subject_ids),
                    request=request,
                )
            report.created += 1
            report.created_users.append({"row": row.row_number, "id": str(user.id), "email": user.email})
        except Conflict:
            report.already_existing += 1
            report.errors.append({"row": row.row_number, "email": data["email"], "errors": ["User already exists."]})
        except APIError as exc:
            report.invalid += 1
            report.errors.append({"row": row.row_number, "email": data["email"],
                                  "errors": [exc.message] + [f"{k}: {v}" for k, v in (exc.details or {}).items()]})

    audit.record(actor, "users.imported", None, {"role": role, **{k: v for k, v in report.as_dict().items() if k in ("total_rows", "created", "already_existing", "invalid")}}, request)
    return report
